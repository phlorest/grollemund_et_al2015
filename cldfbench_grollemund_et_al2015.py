import pathlib
from collections import Counter

from nexus import NexusWriter, NexusReader
from openpyxl import load_workbook
import phlorest


class Dataset(phlorest.Dataset):
    dir = pathlib.Path(__file__).parent
    id = "grollemund_et_al2015"

    def cmd_makecldf(self, args):
        self.init(args)

        with self.nexus_summary() as nex:
            self.add_tree_from_nexus(
                args,
                self.raw_dir / 'grollemund.mcct.trees',
                nex,
                'summary',
                detranslate=True,
            )
        posterior = self.sample(
            self.read_gzipped_text(
                self.raw_dir / 'BP425_M1P_100_cv2_relaxed_YP_runs_1_2_4_5_thinned-fixed.trees.gz'),
            detranslate=True,
            n=100,
            as_nexus=True)

        with self.nexus_posterior() as nex:
            for i, tree in enumerate(posterior.trees.trees, start=1):
                self.add_tree(args, tree, nex, 'posterior-{}'.format(i))

        self.add_data(args, make_nexus(self.raw_dir / 'Grollemund-et-al_Bantu-database_2015.xlsx'))


TEST_DATA = {}
TEST_DATA['arm'] = Counter({
    '1': 260, '10': 2, '11': 1, '12': 1, '13': 9, '14': 8, '15': 2, '16': 2, '17': 1, '18': 4,
    '19': 1,  '2': 8, '20': 1, '21': 64, '22': 12, '23': 1, '24': 1, '25': 3, '26': 3, '27': 2,
    '3': 7, '4': 8, '5': 1, '6': 2, '7': 2, '8': 1, '9': 11, '?': 6
})
TEST_DATA['sun'] = Counter({
    '1': 5, '10': 1, '11': 1, '12': 34, '13': 1, '14': 7, '15': 25, '16': 1, '17': 1, '18': 3,
    '19': 1, '2': 2, '20': 2, '21': 2, '22': 2, '23': 1, '24': 19, '25': 7, '26': 4, '27': 3,
    '28': 4, '29': 1, '3': 1, '30': 4, '31': 1, '32': 1, '33': 46, '34': 1, '35': 1, '36': 1,
    '37': 1, '38': 1, '39': 7, '4': 1, '40': 7, '41': 1, '42': 1, '43': 5, '44': 2, '45': 1,
    '46': 2, '47': 1, '48': 1, '49': 2, '5': 168, '50': 1, '51': 2, '52': 2, '53': 1, '54': 3,
    '55': 1, '56': 1, '57': 1, '58': 2, '59': 6, '6': 1, '60': 1, '61': 1, '62': 1, '63': 2,
    '64': 3, '65': 2, '7': 1, '8': 4, '9': 1, '?': 3
})

TEST_DATA['nose'] = Counter({
    '1': 7, '10': 5, '11': 6, '12': 2, '13': 1, '14': 2, '15': 4, '16': 1, '17': 3, '18': 1,
    '19': 4, '2': 2, '20': 1, '21': 5, '22': 6, '23': 4, '24': 6, '25': 1, '26': 11, '27': 3,
    '28': 33, '29': 65, '3': 1, '30': 3, '31': 5, '32': 12, '33': 11, '34': 1, '35': 2, '36': 1,
    '37': 1, '38': 1, '39': 1, '4': 1, '40': 1, '41': 2, '5': 137, '6': 23, '7': 39, '8': 3,
    '9': 1, '?': 5
})


def test_cognates(cogs, langs):
    for word in TEST_DATA:
        for cogset in TEST_DATA[word]:
            if cogset == '?':
                assert len([l for l in langs if word in langs[l]]) == TEST_DATA[word][cogset]
            else:
                label = (word, int(cogset))
                assert label in cogs, 'missing {}'.format(label)
                assert TEST_DATA[word][cogset] == len(cogs[label])
    return True


def read(filename):
    def s(v):
        return v.strip() if isinstance(v, str) else v

    wb = load_workbook(filename=str(filename), read_only=True)
    header = None
    for i, row in enumerate(wb.worksheets[2].iter_rows(), 1):
        row = [s(cell.value) for cell in row]
        if row[0] in ('Words', 'français'):
            continue
        elif row[0] == 'anglais':
            header = row
        else:
            row = dict(zip(header, row))
            lang = row.pop('anglais').lstrip('*')
            yield (lang, row)


def make_nexus(in_):
    cognates = {}
    languages = {}

    for lang, entries in read(in_):
        languages[lang] = []
        for e in entries:
            if entries[e] == '?':
                languages[lang].append(e)
                continue
            cog = (e, entries[e])
            cognates[cog] = cognates.get(cog, set())
            cognates[cog].add(lang)

    nex = NexusWriter()

    assert test_cognates(cognates, languages)

    state_count = Counter()
    for cog in cognates:
        for lang in languages:
            if cog[0] in languages[lang]:
                state = '?'
            elif lang in cognates[cog]:
                state_count[cog] += 1
                state = '1'
            else:
                state = '0'
            nex.add(lang, "%s_%d" % cog, state)

    for cog in cognates:
        expected = len(cognates[cog])
        assert expected == state_count[cog]
        word, state = cog
        if word in TEST_DATA:
            assert state_count[cog] == TEST_DATA[word][str(state)]

    return NexusReader.from_string(nex.write())
